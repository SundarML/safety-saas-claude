# hira/views.py
import csv
from datetime import date

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.core.exceptions import PermissionDenied
from django.db.models import Count, Q
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone

from .forms import HazardFormSet, HazardRegisterForm
from .models import Hazard, HazardRegister
from .pdf_report import generate_hira_pdf


# ── Guards ────────────────────────────────────────────────────────────────────

def _org(request):
    org = getattr(request, "organization", None)
    if not org:
        raise PermissionDenied
    return org


def _manager_required(request):
    if not (request.user.is_manager or request.user.is_safety_manager):
        raise PermissionDenied


# ── Helpers ───────────────────────────────────────────────────────────────────

def _get_org_users(org):
    from django.contrib.auth import get_user_model
    User = get_user_model()
    return User.objects.filter(organization=org, is_active=True).order_by("full_name", "email")


def _get_compliance_items(org):
    from compliance.models import ComplianceItem
    return ComplianceItem.objects.filter(organization=org).exclude(
        status="not_applicable"
    ).order_by("title")


# ── Dashboard ─────────────────────────────────────────────────────────────────

@login_required
def dashboard(request):
    org = _org(request)
    registers = HazardRegister.objects.filter(organization=org).prefetch_related("hazards")

    # Auto-expire approved registers whose review date has passed
    today = timezone.now().date()
    registers.filter(
        status=HazardRegister.STATUS_APPROVED,
        next_review_date__lt=today,
    ).update(status=HazardRegister.STATUS_EXPIRED)

    registers = HazardRegister.objects.filter(organization=org).prefetch_related("hazards")

    total       = registers.count()
    approved    = registers.filter(status="approved").count()
    draft       = registers.filter(status__in=["draft", "under_review"]).count()
    expired     = registers.filter(status="expired").count()
    review_soon = registers.filter(
        status="approved",
        next_review_date__range=[today, today + timezone.timedelta(days=30)],
    ).count()

    # Count hazards by effective risk level
    all_hazards = Hazard.objects.filter(register__organization=org)
    critical_count = sum(1 for h in all_hazards if h.effective_risk_level == "critical")
    high_count     = sum(1 for h in all_hazards if h.effective_risk_level == "high")
    medium_count   = sum(1 for h in all_hazards if h.effective_risk_level == "medium")
    low_count      = sum(1 for h in all_hazards if h.effective_risk_level == "low")

    # Open actions
    open_actions = all_hazards.filter(action_required=True, action_owner__isnull=False)
    overdue_actions = [h for h in open_actions if h.action_due_date and h.action_due_date < today]

    recent = registers.order_by("-updated_at")[:8]

    return render(request, "hira/dashboard.html", {
        "total":           total,
        "approved":        approved,
        "draft":           draft,
        "expired":         expired,
        "review_soon":     review_soon,
        "critical_count":  critical_count,
        "high_count":      high_count,
        "medium_count":    medium_count,
        "low_count":       low_count,
        "open_actions":    open_actions.count(),
        "overdue_actions": len(overdue_actions),
        "recent":          recent,
    })


# ── Register list ─────────────────────────────────────────────────────────────

@login_required
def register_list(request):
    org = _org(request)
    qs  = HazardRegister.objects.filter(organization=org).prefetch_related("hazards")

    status_filter = request.GET.get("status", "")
    if status_filter:
        qs = qs.filter(status=status_filter)

    return render(request, "hira/register_list.html", {
        "registers":     qs,
        "status_filter": status_filter,
    })


# ── Create ────────────────────────────────────────────────────────────────────

@login_required
def register_create(request):
    org = _org(request)
    _manager_required(request)

    org_users        = _get_org_users(org)
    compliance_items = _get_compliance_items(org)

    def _patch_formset(fs):
        for f in fs.forms:
            f.fields["action_owner"].queryset    = org_users
            f.fields["compliance_item"].queryset = compliance_items

    if request.method == "POST":
        form    = HazardRegisterForm(request.POST)
        formset = HazardFormSet(request.POST, queryset=Hazard.objects.none())
        _patch_formset(formset)

        if form.is_valid() and formset.is_valid():
            register = form.save(commit=False)
            register.organization = org
            register.assessed_by  = request.user
            register.save()

            formset.instance = register
            hazards = formset.save(commit=False)
            for i, h in enumerate(hazards):
                h.register = register
                h.order    = i
                h.save()
            formset.save_m2m()

            messages.success(request, "HIRA register created successfully.")
            return redirect("hira:register_detail", pk=register.pk)
    else:
        form    = HazardRegisterForm(initial={"assessment_date": timezone.now().date()})
        formset = HazardFormSet(queryset=Hazard.objects.none())
        _patch_formset(formset)

    return render(request, "hira/register_form.html", {
        "form":             form,
        "formset":          formset,
        "org_users":        org_users,
        "compliance_items": compliance_items,
        "page_title":       "New HIRA Register",
        "is_edit":          False,
    })


# ── Edit ──────────────────────────────────────────────────────────────────────

@login_required
def register_edit(request, pk):
    org      = _org(request)
    _manager_required(request)
    register         = get_object_or_404(HazardRegister, pk=pk, organization=org)
    org_users        = _get_org_users(org)
    compliance_items = _get_compliance_items(org)

    def _patch_formset(fs):
        for f in fs.forms:
            f.fields["action_owner"].queryset    = org_users
            f.fields["compliance_item"].queryset = compliance_items

    if request.method == "POST":
        form    = HazardRegisterForm(request.POST, instance=register)
        formset = HazardFormSet(request.POST, instance=register)
        _patch_formset(formset)

        if form.is_valid() and formset.is_valid():
            form.save()
            hazards = formset.save(commit=False)
            for i, h in enumerate(hazards):
                h.register = register
                h.order    = i
                h.save()
            for h in formset.deleted_objects:
                h.delete()
            formset.save_m2m()

            messages.success(request, "HIRA register updated.")
            return redirect("hira:register_detail", pk=register.pk)
    else:
        form    = HazardRegisterForm(instance=register)
        formset = HazardFormSet(instance=register)
        _patch_formset(formset)

    return render(request, "hira/register_form.html", {
        "form":             form,
        "formset":          formset,
        "register":         register,
        "org_users":        org_users,
        "compliance_items": compliance_items,
        "page_title":       f"Edit — {register.title}",
        "is_edit":          True,
    })


# ── Detail ────────────────────────────────────────────────────────────────────

@login_required
def register_detail(request, pk):
    org      = _org(request)
    register = get_object_or_404(
        HazardRegister.objects.prefetch_related(
            "hazards__action_owner",
            "hazards__linked_observations",
            "hazards__compliance_item",
        ),
        pk=pk, organization=org,
    )
    hazards = list(register.hazards.all())
    risk_counts = {
        "critical": sum(1 for h in hazards if h.effective_risk_level == "critical"),
        "high":     sum(1 for h in hazards if h.effective_risk_level == "high"),
        "medium":   sum(1 for h in hazards if h.effective_risk_level == "medium"),
        "low":      sum(1 for h in hazards if h.effective_risk_level == "low"),
    }
    return render(request, "hira/register_detail.html", {
        "register":    register,
        "risk_counts": risk_counts,
        "today":       timezone.now().date(),
    })


# ── Approve ───────────────────────────────────────────────────────────────────

@login_required
def register_approve(request, pk):
    org = _org(request)
    _manager_required(request)
    register = get_object_or_404(HazardRegister, pk=pk, organization=org)

    if request.method == "POST":
        action = request.POST.get("action")
        if action == "submit" and register.status == HazardRegister.STATUS_DRAFT:
            register.status = HazardRegister.STATUS_UNDER_REVIEW
            register.save()
            messages.success(request, "Register submitted for review.")
        elif action == "approve" and register.status == HazardRegister.STATUS_UNDER_REVIEW:
            register.status      = HazardRegister.STATUS_APPROVED
            register.approved_by = request.user
            register.approved_at = timezone.now()
            register.save()
            messages.success(request, "Register approved.")
        elif action == "revert":
            register.status = HazardRegister.STATUS_DRAFT
            register.save()
            messages.info(request, "Register reverted to Draft.")

    return redirect("hira:register_detail", pk=register.pk)


# ── Delete ────────────────────────────────────────────────────────────────────

@login_required
def register_delete(request, pk):
    org = _org(request)
    _manager_required(request)
    register = get_object_or_404(HazardRegister, pk=pk, organization=org)

    if request.method == "POST":
        register.delete()
        messages.success(request, "Register deleted.")
        return redirect("hira:register_list")

    return render(request, "hira/register_confirm_delete.html", {"register": register})


# ── PDF ───────────────────────────────────────────────────────────────────────

@login_required
def register_pdf(request, pk):
    org      = _org(request)
    register = get_object_or_404(
        HazardRegister.objects.prefetch_related("hazards__action_owner"),
        pk=pk, organization=org,
    )
    pdf_bytes = generate_hira_pdf(register)
    filename  = f"HIRA-{register.pk:04d}-Rev{register.revision_no}.pdf"
    response  = HttpResponse(pdf_bytes, content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


# ── Export helpers ────────────────────────────────────────────────────────────

EXPORT_HEADERS = [
    # Register fields (11)
    "Register ID", "Register Title", "Activity / Work Area", "Location",
    "Assessment Date", "Next Review Date", "Status", "Revision No.",
    "Assessed By", "Approved By", "Approved On",
    # Hazard fields (5)
    "Hazard #", "Category", "Hazard Description", "Potential Harm",
    "Who Might Be Harmed",
    # Initial risk (4)
    "Initial Likelihood", "Initial Severity", "Initial Risk Score", "Initial Risk Level",
    # Controls (2)
    "Primary Control Type", "Controls Description",
    # Residual risk (4)
    "Residual Likelihood", "Residual Severity", "Residual Risk Score", "Residual Risk Level",
    # Action (3)
    "Action Required", "Action Owner", "Action Due Date",
]


def _hazard_row(register, hazard, hazard_num):
    """Return a flat list of cell values for one hazard row."""
    reg_fields = [
        register.pk,
        register.title,
        register.activity,
        register.location_text or "",
        register.assessment_date.strftime("%Y-%m-%d"),
        register.next_review_date.strftime("%Y-%m-%d") if register.next_review_date else "",
        register.get_status_display(),
        register.revision_no,
        register.assessed_by.get_full_name() if register.assessed_by else "",
        register.approved_by.get_full_name() if register.approved_by else "",
        timezone.localtime(register.approved_at).strftime("%Y-%m-%d %H:%M") if register.approved_at else "",
    ]
    if hazard is None:
        return reg_fields + [""] * (len(EXPORT_HEADERS) - 11)

    return reg_fields + [
        hazard_num,
        hazard.get_category_display(),
        hazard.hazard_description,
        hazard.potential_harm,
        hazard.get_who_might_be_harmed_display(),
        hazard.initial_likelihood,
        hazard.initial_severity,
        hazard.initial_risk_score,
        hazard.initial_risk_level.title() if hazard.initial_risk_level else "",
        hazard.get_primary_control_type_display(),
        hazard.controls_description,
        hazard.residual_likelihood or "",
        hazard.residual_severity or "",
        hazard.residual_risk_score or "",
        hazard.residual_risk_level.title() if hazard.residual_risk_level else "",
        "Yes" if hazard.action_required else "No",
        hazard.action_owner.get_full_name() if hazard.action_owner else "",
        hazard.action_due_date.strftime("%Y-%m-%d") if hazard.action_due_date else "",
    ]


def _export_queryset(org):
    return (
        HazardRegister.objects
        .filter(organization=org)
        .prefetch_related("hazards__action_owner")
        .select_related("assessed_by", "approved_by")
        .order_by("-assessment_date", "pk")
    )


# ── CSV export ────────────────────────────────────────────────────────────────

@login_required
def export_csv(request):
    org = _org(request)
    _manager_required(request)

    response = HttpResponse(content_type="text/csv")
    response["Content-Disposition"] = (
        f'attachment; filename="HIRA-export-{date.today()}.csv"'
    )
    writer = csv.writer(response)
    writer.writerow(EXPORT_HEADERS)

    for register in _export_queryset(org):
        hazards = list(register.hazards.all())
        if hazards:
            for i, h in enumerate(hazards, 1):
                writer.writerow(_hazard_row(register, h, i))
        else:
            writer.writerow(_hazard_row(register, None, None))

    return response


# ── Excel export ──────────────────────────────────────────────────────────────

@login_required
def export_excel(request):
    org = _org(request)
    _manager_required(request)

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from core.logo_utils import get_logo_for_excel

    RISK_FILLS = {
        "critical": ("FEE2E2", "991B1B"),
        "high":     ("FFEDD5", "9A3412"),
        "medium":   ("FEF9C3", "854D0E"),
        "low":      ("D1FAE5", "065F46"),
    }

    wb = Workbook()
    ws = wb.active
    ws.title = "HIRA Register"

    # ── Logo + org header ─────────────────────────────────────────────────────
    xl_logo = get_logo_for_excel(org)
    DATA_START_ROW = 1

    if xl_logo:
        xl_logo.anchor = "A1"
        ws.add_image(xl_logo)
        for r in range(1, 4):
            ws.row_dimensions[r].height = 14
        ws["B1"] = org.name
        ws["B1"].font = Font(bold=True, size=13, color="0E1729")
        ws["B2"] = "Hazard Identification & Risk Assessment (HIRA) Export"
        ws["B2"].font = Font(size=9, color="64748B")
        ws["B3"] = f"Exported: {date.today().strftime('%d %b %Y')}  |  ISO 45001 Aligned"
        ws["B3"].font = Font(size=9, color="94A3B8")
        ws.row_dimensions[4].height = 6
        DATA_START_ROW = 5

    # ── Column headers ────────────────────────────────────────────────────────
    thin   = Side(style="thin", color="DEE2E6")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Group colour definitions: (bg hex, fg hex) per column index (1-based)
    group_colors = {}
    for i in range(1, 12):   group_colors[i] = ("1E3A5F", "FFFFFF")  # Register
    for i in range(12, 17):  group_colors[i] = ("334155", "FFFFFF")  # Hazard info
    for i in range(17, 21):  group_colors[i] = ("7F1D1D", "FFFFFF")  # Initial risk
    for i in range(21, 23):  group_colors[i] = ("14532D", "FFFFFF")  # Controls
    for i in range(23, 27):  group_colors[i] = ("7C2D12", "FFFFFF")  # Residual risk
    for i in range(27, 30):  group_colors[i] = ("78350F", "FFFFFF")  # Action

    for col_idx, header in enumerate(EXPORT_HEADERS, start=1):
        bg, fg = group_colors.get(col_idx, ("0E1729", "FFFFFF"))
        cell = ws.cell(row=DATA_START_ROW, column=col_idx, value=header)
        cell.fill = PatternFill("solid", fgColor=bg)
        cell.font = Font(bold=True, color=fg, size=9)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
    ws.row_dimensions[DATA_START_ROW].height = 30

    # ── Data rows ─────────────────────────────────────────────────────────────
    row_num = DATA_START_ROW + 1
    alt_fill   = PatternFill("solid", fgColor="F8F9FB")
    white_fill = PatternFill("solid", fgColor="FFFFFF")

    for reg_idx, register in enumerate(_export_queryset(org)):
        hazards = list(register.hazards.all())
        rows_to_write = [(i + 1, h) for i, h in enumerate(hazards)] if hazards else [(None, None)]
        fill = alt_fill if reg_idx % 2 == 0 else white_fill

        for h_num, hazard in rows_to_write:
            row_data = _hazard_row(register, hazard, h_num)

            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_num, column=col_idx, value=value)
                cell.fill = fill
                cell.border = border
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                cell.font = Font(size=9)

            # Colour risk level cells
            if hazard:
                init_level = hazard.initial_risk_level
                if init_level in RISK_FILLS:
                    bg, fg = RISK_FILLS[init_level]
                    c = ws.cell(row=row_num, column=20)
                    c.fill = PatternFill("solid", fgColor=bg)
                    c.font = Font(bold=True, color=fg, size=9)

                res_level = hazard.residual_risk_level
                if res_level and res_level in RISK_FILLS:
                    bg, fg = RISK_FILLS[res_level]
                    c = ws.cell(row=row_num, column=26)
                    c.fill = PatternFill("solid", fgColor=bg)
                    c.font = Font(bold=True, color=fg, size=9)

                if hazard.action_required:
                    c = ws.cell(row=row_num, column=27)
                    c.fill = PatternFill("solid", fgColor="FFF7ED")
                    c.font = Font(bold=True, color="EA580C", size=9)

            ws.row_dimensions[row_num].height = 36
            row_num += 1

    # ── Column widths ─────────────────────────────────────────────────────────
    col_widths = [
        8,  30, 28, 18, 14, 14, 14,  8, 22, 22, 18,   # Register (11)
        6,  18, 40, 35, 16,                             # Hazard info (5)
        12, 12, 10, 12,                                 # Initial risk (4)
        20, 45,                                         # Controls (2)
        14, 14, 12, 14,                                 # Residual risk (4)
        12, 22, 14,                                     # Action (3)
    ]
    for i, w in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Freeze header row + register columns
    ws.freeze_panes = ws.cell(row=DATA_START_ROW + 1, column=12)

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = (
        f'attachment; filename="HIRA-export-{date.today()}.xlsx"'
    )
    wb.save(response)
    return response


# ── Risk Matrix ───────────────────────────────────────────────────────────────

@login_required
def risk_matrix(request):
    org = _org(request)
    all_hazards = (
        Hazard.objects
        .filter(register__organization=org)
        .select_related("register")
    )

    # Build matrix[likelihood][severity] = list of (hazard, register)
    matrix = {l: {s: [] for s in range(1, 6)} for l in range(1, 6)}
    for hazard in all_hazards:
        if hazard.residual_likelihood and hazard.residual_severity:
            l, s = hazard.residual_likelihood, hazard.residual_severity
        else:
            l, s = hazard.initial_likelihood, hazard.initial_severity
        matrix[l][s].append(hazard)

    from .models import compute_risk_level

    LIKE_LABELS = {5: "Almost Certain", 4: "Likely", 3: "Possible", 2: "Unlikely", 1: "Rare"}
    SEV_LABELS  = {1: "Insignificant", 2: "Minor", 3: "Moderate", 4: "Major", 5: "Catastrophic"}

    # For the selected cell detail (optional query param)
    sel_l = sel_s = None
    raw_l, raw_s  = request.GET.get("l"), request.GET.get("s")
    if raw_l and raw_s:
        try:
            sl, ss = int(raw_l), int(raw_s)
            if 1 <= sl <= 5 and 1 <= ss <= 5:
                sel_l, sel_s = sl, ss
        except ValueError:
            pass

    # Build rows list (likelihood 5→1, top to bottom)
    cell_rows = []
    for l in range(5, 0, -1):
        cells = []
        for s in range(1, 6):
            score   = l * s
            level   = compute_risk_level(score)
            hazards = matrix[l][s]
            cells.append({
                "l": l, "s": s, "score": score, "level": level,
                "count": len(hazards), "hazards": hazards,
                "selected": (l == sel_l and s == sel_s),
            })
        cell_rows.append({"l": l, "l_label": LIKE_LABELS[l], "cells": cells})

    sev_headers = [{"s": s, "label": SEV_LABELS[s]} for s in range(1, 6)]

    selected_hazards = None
    selected_score   = None
    if sel_l and sel_s:
        selected_hazards = matrix[sel_l][sel_s]
        selected_score   = sel_l * sel_s

    return render(request, "hira/risk_matrix.html", {
        "cell_rows":        cell_rows,
        "sev_headers":      sev_headers,
        "selected_l":       sel_l,
        "selected_s":       sel_s,
        "selected_hazards": selected_hazards,
        "selected_score":   selected_score,
    })


# ── Observation ↔ Hazard linking ──────────────────────────────────────────────

@login_required
def link_observation(request, hazard_pk):
    """Link an existing observation to a hazard (managers only)."""
    org = _org(request)
    _manager_required(request)
    hazard = get_object_or_404(Hazard, pk=hazard_pk, register__organization=org)

    from observations.models import Observation
    already_linked = hazard.linked_observations.values_list("pk", flat=True)
    available_obs  = (
        Observation.objects
        .filter(organization=org)
        .exclude(pk__in=already_linked)
        .order_by("-date_observed")[:200]
    )

    if request.method == "POST":
        obs_pk = request.POST.get("observation_pk")
        if obs_pk:
            obs = get_object_or_404(Observation, pk=obs_pk, organization=org)
            hazard.linked_observations.add(obs)
            messages.success(request, f"Observation #{obs.pk} linked to this hazard.")
        return redirect("hira:register_detail", pk=hazard.register_id)

    return render(request, "hira/link_observation.html", {
        "hazard":         hazard,
        "available_obs":  available_obs,
    })


@login_required
def unlink_observation(request, hazard_pk, obs_pk):
    """Remove observation link from a hazard (managers only, POST only)."""
    org = _org(request)
    _manager_required(request)
    hazard = get_object_or_404(Hazard, pk=hazard_pk, register__organization=org)

    if request.method == "POST":
        from observations.models import Observation
        obs = get_object_or_404(Observation, pk=obs_pk, organization=org)
        hazard.linked_observations.remove(obs)
        messages.success(request, "Observation unlinked.")

    return redirect("hira:register_detail", pk=hazard.register_id)
